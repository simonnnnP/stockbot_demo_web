# excel.py  – funciones mínimas para leer y actualizar el archivo stock.xlsx
import openpyxl

EXCEL_PATH = "stock.xlsx"   # debe existir en la misma carpeta

def leer_stock():
    """
    Devuelve una lista de diccionarios con los datos de la hoja 'Stock'.
    Columnas esperadas en la fila 1: SKU | Descripción | Cantidad | Unidad | Ubicación
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Stock"]
    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        datos.append({
            "sku": row[0],
            "desc": row[1],
            "qty": row[2],
            "unit": row[3],
            "loc": row[4],
        })
    wb.close()
    return datos

def actualizar_stock(sku, cantidad, operacion):
    """
    Suma o resta 'cantidad' al SKU indicado.
    operacion = 'ingreso' para sumar • 'egreso' para restar (y valida stock suficiente)
    Lanza ValueError si no hay stock.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Stock"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sku:
            actual = row[2].value or 0
            if operacion == "ingreso":
                row[2].value = actual + cantidad
            elif operacion == "egreso":
                if actual < cantidad:
                    wb.close()
                    raise ValueError("Stock insuficiente")
                row[2].value = actual - cantidad
            break
    else:
        wb.close()
        raise ValueError(f"SKU {sku} no encontrado")
    wb.save(EXCEL_PATH)
    wb.close()
